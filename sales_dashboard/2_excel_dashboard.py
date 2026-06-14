"""
Sales Performance Dashboard — Excel Workbook Builder
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os

INPUT  =  r"C:\Users\HP\Downloads\sales_dashboard_project\sales_dashboard\ecommerce_sales_dataset.csv"
OUTPUT = "outputs/Sales_Dashboard.xlsx"
os.makedirs("outputs", exist_ok=True)

# ── Styles ────────────────────────────────────────────────────────────────────
DARK_BLUE  = "1B3A6B"
MID_BLUE   = "378ADD"
LIGHT_BLUE = "E6F1FB"
GREEN      = "1D9E75"
ORANGE     = "D85A30"
PURPLE     = "7F77DD"
AMBER      = "BA7517"
WHITE      = "FFFFFF"
GRAY_BG    = "F4F6F9"
GRAY_BORDER= "D0D5DD"

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="1A1A1A"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def border(color=GRAY_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def money_fmt():  return '#,##0.00'
def pct_fmt():    return '0.00%'
def int_fmt():    return '#,##0'

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def style_header_row(ws, row, start_col, end_col, bg=DARK_BLUE, font_color=WHITE, height=24):
    ws.row_dimensions[row].height = height
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(bg)
        cell.font = hdr_font(color=font_color)
        cell.alignment = center()
        cell.border = border()

def style_data_row(ws, row, start_col, end_col, bg=WHITE):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(bg)
        cell.font = body_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border()

# ── Load data ─────────────────────────────────────────────────────────────────
df = pd.read_csv(INPUT)
df = df[df["Order_Status"] != "Cancelled"].copy()

wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — KPI Summary
# ══════════════════════════════════════════════════════════════════════════════
ws_kpi = wb.create_sheet("📊 KPI Summary")
ws_kpi.sheet_view.showGridLines = False
ws_kpi.sheet_view.zoomScale = 90

# Title banner
ws_kpi.merge_cells("A1:H1")
ws_kpi["A1"] = "SALES PERFORMANCE DASHBOARD"
ws_kpi["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
ws_kpi["A1"].fill = fill(DARK_BLUE)
ws_kpi["A1"].alignment = center()
ws_kpi.row_dimensions[1].height = 38

ws_kpi.merge_cells("A2:H2")
ws_kpi["A2"] = "E-Commerce Sales Analysis  |  2021 – 2024  |  4 Regions  |  5 Categories"
ws_kpi["A2"].font = Font(name="Arial", size=10, color="666666")
ws_kpi["A2"].fill = fill(LIGHT_BLUE)
ws_kpi["A2"].alignment = center()
ws_kpi.row_dimensions[2].height = 20

ws_kpi.row_dimensions[3].height = 14

# KPI cards (row 4–7)
kpis = [
    ("Total Revenue",     f"=${df['Revenue'].sum():,.2f}",    MID_BLUE,  "💰"),
    ("Total Profit",      f"=${df['Profit'].sum():,.2f}",     GREEN,     "📈"),
    ("Total Orders",      f"{len(df):,}",                     PURPLE,    "🛒"),
    ("Avg Profit Margin", f"{df['Profit_Margin_%'].mean():.2f}%", ORANGE, "🎯"),
]

cols = [2, 4, 6, 8]
for i, (label, value, color, icon) in enumerate(kpis):
    c = cols[i]
    ws_kpi.merge_cells(start_row=4, start_column=c, end_row=4, end_column=c)
    ws_kpi.merge_cells(start_row=5, start_column=c, end_row=5, end_column=c)
    ws_kpi.merge_cells(start_row=6, start_column=c, end_row=6, end_column=c)

    lbl_cell = ws_kpi.cell(row=4, column=c, value=f"{icon}  {label}")
    lbl_cell.font = Font(name="Arial", size=9, color=WHITE)
    lbl_cell.fill = fill(color)
    lbl_cell.alignment = center()
    ws_kpi.row_dimensions[4].height = 20

    val_cell = ws_kpi.cell(row=5, column=c, value=value)
    val_cell.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    val_cell.fill = fill(color)
    val_cell.alignment = center()
    ws_kpi.row_dimensions[5].height = 28

    ws_kpi.cell(row=6, column=c).fill = fill(color)
    ws_kpi.row_dimensions[6].height = 6

ws_kpi.row_dimensions[7].height = 14

# Section headers
def section_header(ws, row, col, text):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+4)
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Arial", size=10, bold=True, color=DARK_BLUE)
    c.fill = fill(LIGHT_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20

# Region breakdown
section_header(ws_kpi, 8, 1, "Revenue by Region")
region_data = df.groupby("Region")[["Revenue","Profit","Profit_Margin_%"]].agg(
    {"Revenue":"sum","Profit":"sum","Profit_Margin_%":"mean"}).sort_values("Revenue",ascending=False)
region_data.columns = ["Revenue ($)","Profit ($)","Avg Margin (%)"]
region_data = region_data.reset_index()

headers = ["Region","Revenue ($)","Profit ($)","Avg Margin (%)","% of Total Revenue"]
for j, h in enumerate(headers):
    c = ws_kpi.cell(row=9, column=j+1, value=h)
    c.font = hdr_font(size=9, color=WHITE)
    c.fill = fill(DARK_BLUE)
    c.alignment = center()
    c.border = border()
    ws_kpi.row_dimensions[9].height = 20

total_rev = df["Revenue"].sum()
for i, row_data in region_data.iterrows():
    r = 10 + i
    bg = GRAY_BG if i % 2 == 0 else WHITE
    vals = [row_data["Region"], row_data["Revenue ($)"],
            row_data["Profit ($)"], row_data["Avg Margin (%)"]/100,
            row_data["Revenue ($)"]/total_rev]
    fmts = [None, money_fmt(), money_fmt(), pct_fmt(), pct_fmt()]
    for j, (v, fmt) in enumerate(zip(vals, fmts)):
        c = ws_kpi.cell(row=r, column=j+1, value=v)
        c.font = body_font()
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border()
        if fmt:
            c.number_format = fmt
    ws_kpi.row_dimensions[r].height = 18

# Totals row
r = 10 + len(region_data)
ws_kpi.cell(row=r, column=1, value="TOTAL").font = hdr_font(size=9, color=WHITE)
ws_kpi.cell(row=r, column=1).fill = fill(DARK_BLUE)
ws_kpi.cell(row=r, column=1).alignment = center()
ws_kpi.cell(row=r, column=2, value=f"=SUM(B10:B{r-1})").number_format = money_fmt()
ws_kpi.cell(row=r, column=3, value=f"=SUM(C10:C{r-1})").number_format = money_fmt()
for col in [2,3,4,5]:
    c = ws_kpi.cell(row=r, column=col)
    c.font = hdr_font(size=9, color=WHITE)
    c.fill = fill(DARK_BLUE)
    c.alignment = center()
    c.border = border()
ws_kpi.row_dimensions[r].height = 18

for col, width in [(1,20),(2,18),(3,16),(4,16),(5,20),(6,14),(7,14),(8,14)]:
    set_col_width(ws_kpi, col, width)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Regional Analysis
# ══════════════════════════════════════════════════════════════════════════════
ws_reg = wb.create_sheet("🌍 Regional Analysis")
ws_reg.sheet_view.showGridLines = False

ws_reg.merge_cells("A1:G1")
ws_reg["A1"] = "REGIONAL SALES ANALYSIS"
ws_reg["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws_reg["A1"].fill = fill(DARK_BLUE)
ws_reg["A1"].alignment = center()
ws_reg.row_dimensions[1].height = 32

reg_quarterly = df.groupby(["Region","Year","Quarter"])["Revenue"].sum().reset_index()
reg_cat = df.groupby(["Region","Category"])["Revenue"].sum().reset_index()

# Region × Year pivot
pivot_ry = df.pivot_table(index="Region", columns="Year", values="Revenue", aggfunc="sum")
pivot_ry.columns = [str(c) for c in pivot_ry.columns]
pivot_ry = pivot_ry.reset_index()

ws_reg.cell(row=3, column=1, value="Revenue by Region & Year ($)").font = \
    Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
ws_reg.row_dimensions[3].height = 22

cols_ry = ["Region"] + [str(y) for y in sorted(df["Year"].unique())] + ["Total"]
for j, h in enumerate(cols_ry):
    c = ws_reg.cell(row=4, column=j+1, value=h)
    c.font = hdr_font(size=9)
    c.fill = fill(DARK_BLUE)
    c.alignment = center()
    c.border = border()
ws_reg.row_dimensions[4].height = 20

for i, row_data in pivot_ry.iterrows():
    r = 5 + i
    bg = GRAY_BG if i % 2 == 0 else WHITE
    for j, h in enumerate(cols_ry):
        if h == "Total":
            val = f"=SUM(B{r}:E{r})"
        elif h == "Region":
            val = row_data["Region"]
        else:
            val = row_data.get(h, 0)
        c = ws_reg.cell(row=r, column=j+1, value=val)
        c.font = body_font()
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border()
        if h != "Region":
            c.number_format = money_fmt()
    ws_reg.row_dimensions[r].height = 18

# Bar chart — Region × Year
chart_r = BarChart()
chart_r.type = "col"
chart_r.grouping = "clustered"
chart_r.title = "Revenue by Region & Year"
chart_r.y_axis.title = "Revenue ($)"
chart_r.style = 10
chart_r.width = 18
chart_r.height = 11

data_ref = Reference(ws_reg, min_col=2, max_col=5, min_row=4, max_row=8)
cats_ref = Reference(ws_reg, min_col=1, min_row=5, max_row=8)
chart_r.add_data(data_ref, titles_from_data=True)
chart_r.set_categories(cats_ref)
ws_reg.add_chart(chart_r, "A11")

for col, width in [(1,18),(2,16),(3,16),(4,16),(5,16),(6,16)]:
    set_col_width(ws_reg, col, width)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Product Analysis
# ══════════════════════════════════════════════════════════════════════════════
ws_prod = wb.create_sheet("📦 Product Analysis")
ws_prod.sheet_view.showGridLines = False

ws_prod.merge_cells("A1:G1")
ws_prod["A1"] = "PRODUCT & CATEGORY ANALYSIS"
ws_prod["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws_prod["A1"].fill = fill(DARK_BLUE)
ws_prod["A1"].alignment = center()
ws_prod.row_dimensions[1].height = 32

cat_summary = (
    df.groupby("Category")
    .agg(Revenue=("Revenue","sum"), Profit=("Profit","sum"),
         Orders=("Order_ID","count"), Margin=("Profit_Margin_%","mean"))
    .sort_values("Revenue", ascending=False)
    .reset_index()
)

ws_prod.cell(row=3, column=1, value="Category Performance").font = \
    Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
ws_prod.row_dimensions[3].height = 22

cat_cols = ["Category","Revenue ($)","Profit ($)","Orders","Avg Margin (%)","Profit %"]
for j, h in enumerate(cat_cols):
    c = ws_prod.cell(row=4, column=j+1, value=h)
    c.font = hdr_font(size=9)
    c.fill = fill(DARK_BLUE)
    c.alignment = center()
    c.border = border()
ws_prod.row_dimensions[4].height = 20

for i, row_data in cat_summary.iterrows():
    r = 5 + i
    bg = GRAY_BG if i % 2 == 0 else WHITE
    vals = [row_data["Category"], row_data["Revenue"], row_data["Profit"],
            row_data["Orders"], row_data["Margin"]/100,
            row_data["Profit"]/row_data["Revenue"]]
    fmts = [None, money_fmt(), money_fmt(), int_fmt(), pct_fmt(), pct_fmt()]
    for j, (v, fmt) in enumerate(zip(vals, fmts)):
        c = ws_prod.cell(row=r, column=j+1, value=v)
        c.font = body_font()
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border()
        if fmt: c.number_format = fmt
    ws_prod.row_dimensions[r].height = 18

# Top 10 products
top10 = (df.groupby("Product_Name")["Revenue"].sum()
         .sort_values(ascending=False).head(10).reset_index())
top10.columns = ["Product","Revenue ($)"]

ws_prod.cell(row=3, column=9, value="Top 10 Products by Revenue").font = \
    Font(name="Arial", size=11, bold=True, color=DARK_BLUE)

for j, h in enumerate(["Product","Revenue ($)"]):
    c = ws_prod.cell(row=4, column=9+j, value=h)
    c.font = hdr_font(size=9)
    c.fill = fill(DARK_BLUE)
    c.alignment = center()
    c.border = border()

for i, row_data in top10.iterrows():
    r = 5 + i
    bg = GRAY_BG if i % 2 == 0 else WHITE
    for j, (v, fmt) in enumerate(zip([row_data["Product"], row_data["Revenue ($)"]],
                                      [None, money_fmt()])):
        c = ws_prod.cell(row=r, column=9+j, value=v)
        c.font = body_font()
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border()
        if fmt: c.number_format = fmt
    ws_prod.row_dimensions[r].height = 18

# Bar chart — category revenue
chart_c = BarChart()
chart_c.type = "bar"
chart_c.title = "Revenue by Category"
chart_c.y_axis.title = "Category"
chart_c.x_axis.title = "Revenue ($)"
chart_c.style = 10
chart_c.width = 16
chart_c.height = 10
data_c = Reference(ws_prod, min_col=2, max_col=2, min_row=4, max_row=9)
cats_c = Reference(ws_prod, min_col=1, min_row=5, max_row=9)
chart_c.add_data(data_c, titles_from_data=True)
chart_c.set_categories(cats_c)
ws_prod.add_chart(chart_c, "A12")

for col, width in [(1,20),(2,16),(3,14),(4,10),(5,14),(6,12),(7,10),(8,4),(9,26),(10,16)]:
    set_col_width(ws_prod, col, width)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Time Series
# ══════════════════════════════════════════════════════════════════════════════
ws_time = wb.create_sheet("📅 Time Series")
ws_time.sheet_view.showGridLines = False

ws_time.merge_cells("A1:H1")
ws_time["A1"] = "TIME SERIES ANALYSIS"
ws_time["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
ws_time["A1"].fill = fill(DARK_BLUE)
ws_time["A1"].alignment = center()
ws_time.row_dimensions[1].height = 32

quarterly = (
    df.groupby(["Year","Quarter"])[["Revenue","Profit"]]
    .sum().reset_index().sort_values(["Year","Quarter"])
)
quarterly["Label"] = quarterly["Year"].astype(str) + " " + quarterly["Quarter"]

ws_time.cell(row=3, column=1, value="Quarterly Revenue & Profit").font = \
    Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
ws_time.row_dimensions[3].height = 22

for j, h in enumerate(["Period","Revenue ($)","Profit ($)","Profit Margin (%)"]):
    c = ws_time.cell(row=4, column=j+1, value=h)
    c.font = hdr_font(size=9)
    c.fill = fill(DARK_BLUE)
    c.alignment = center()
    c.border = border()
ws_time.row_dimensions[4].height = 20

for i, row_data in quarterly.iterrows():
    r = 5 + i
    bg = GRAY_BG if i % 2 == 0 else WHITE
    margin_formula = f"=C{r}/B{r}"
    vals = [row_data["Label"], row_data["Revenue"], row_data["Profit"], margin_formula]
    fmts = [None, money_fmt(), money_fmt(), pct_fmt()]
    for j, (v, fmt) in enumerate(zip(vals, fmts)):
        c = ws_time.cell(row=r, column=j+1, value=v)
        c.font = body_font()
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border()
        if fmt: c.number_format = fmt
    ws_time.row_dimensions[r].height = 18

last_r = 5 + len(quarterly) - 1

# Line chart — quarterly trend
chart_t = LineChart()
chart_t.title = "Quarterly Revenue & Profit Trend"
chart_t.y_axis.title = "Amount ($)"
chart_t.style = 10
chart_t.width = 22
chart_t.height = 12

data_t = Reference(ws_time, min_col=2, max_col=3, min_row=4, max_row=last_r)
cats_t = Reference(ws_time, min_col=1, min_row=5, max_row=last_r)
chart_t.add_data(data_t, titles_from_data=True)
chart_t.set_categories(cats_t)
chart_t.series[0].graphicalProperties.line.solidFill = MID_BLUE
chart_t.series[1].graphicalProperties.line.solidFill = GREEN
ws_time.add_chart(chart_t, "F3")

for col, width in [(1,14),(2,16),(3,14),(4,16)]:
    set_col_width(ws_time, col, width)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Raw Data
# ══════════════════════════════════════════════════════════════════════════════
ws_raw = wb.create_sheet("📋 Raw Data")
ws_raw.sheet_view.showGridLines = True

ws_raw.merge_cells("A1:Z1")
ws_raw["A1"] = "RAW DATA — ecommerce_sales_dataset.csv"
ws_raw["A1"].font = Font(name="Arial", size=11, bold=True, color=WHITE)
ws_raw["A1"].fill = fill(DARK_BLUE)
ws_raw["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws_raw.row_dimensions[1].height = 24

for j, col_name in enumerate(df.columns):
    c = ws_raw.cell(row=2, column=j+1, value=col_name)
    c.font = hdr_font(size=9)
    c.fill = fill(MID_BLUE)
    c.alignment = center()
    c.border = border()
ws_raw.row_dimensions[2].height = 20

for i, row_data in df.head(5000).iterrows():
    r = 3 + i
    bg = GRAY_BG if i % 2 == 0 else WHITE
    for j, val in enumerate(row_data):
        c = ws_raw.cell(row=r, column=j+1, value=val)
        c.font = body_font(size=9)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_raw.row_dimensions[r].height = 16

ws_raw.auto_filter.ref = f"A2:{get_column_letter(len(df.columns))}2"
ws_raw.freeze_panes = "A3"

for col in range(1, len(df.columns)+1):
    set_col_width(ws_raw, col, 16)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print("Sheets: KPI Summary | Regional Analysis | Product Analysis | Time Series | Raw Data")
